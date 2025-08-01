from pathlib import Path
from typing import Union

import openpyxl


def insert_image_to_cell(ws: openpyxl.Workbook.worksheets,
                         img_path: Union[str, Path],
                         col_idx: int,
                         row_idx: int,
                         scale: float = 1.0):
    """Insert a scaled image to the selected cell of the given worksheet.

    Args:
        ws (openpyxl.Workbook.worksheets): Worksheet
        img_path (Union[str, Path]): Path to the image
        col_idx (int): Column index of the cell
        row_idx (int): Row index of the cell
        scale (float, optional): Scaling factor of the image. Defaults to 1.0.

    Raises:
        ValueError: If the image does not exist.
        TypeError: If cell indices are not integer.
        TypeError: If the scale of the image is not a float.
    """
    img_path = Path(img_path)
    if not img_path.exists():
        raise ValueError('File "{}" does not exist.'.format(img_path))
    if not isinstance(col_idx, int) or not isinstance(row_idx, int):
        raise TypeError('Column and row indices should be integer.')
    if not isinstance(scale, float):
        raise TypeError('scale should be a float number.')
    
    img = openpyxl.drawing.image.Image(img_path)
    img.height = img.height * scale
    img.width = img.width * scale
    col_str = openpyxl.utils.get_column_letter(col_idx)

    ws.row_dimensions[row_idx].height = 8
    ws.row_dimensions[row_idx].height = max(
        img.height * 0.75, ws.row_dimensions[row_idx].height)

    ws.column_dimensions[col_str].width = 15
    ws.column_dimensions[col_str].width = max(
        img.width * 0.128, ws.column_dimensions[col_str].width)

    cell_position = ws.cell(row=row_idx, column=col_idx).coordinate
    img.anchor = cell_position
    ws.add_image(img)


if __name__ == '__main__':
    output = 'example_lenna.xlsx'
    img_path = 'lenna.png'
    col = 2
    row = 2

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    insert_image_to_cell(ws, img_path, col, row, 0.5)

    wb.save(output)
